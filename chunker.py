"""
CMS TWiki GFM Chunker for RAG Pipeline
=======================================
Chunks GFM Markdown files (output of twiki2gfm) into semantic units
for embedding with e5-large-v2 (512 token limit).

Three phases:
  1. Filter  - skip files with insufficient content
  2. Clean   - remove HTML/TWiki artifacts, extract metadata
  3. Chunk   - split by headings, Q&A patterns, or paragraphs

Usage:
  python chunker.py <input_dir> <output_file.jsonl>
  python chunker.py <input_dir> <output_file.jsonl> --analyze
  python chunker.py --sample <file.md>

Author: Buruj Kiahamy (CERN/KAUST Academy)
"""

import os
import re
import json
import sys
import argparse
from pathlib import Path
from dataclasses import dataclass, field, asdict


# --- Configuration -----------------------------------------------------------
MAX_CHUNK_TOKENS   = 400    # target max tokens per chunk (e5-large-v2 limit = 512)
OVERLAP_TOKENS     = 50     # overlap between consecutive chunks
MIN_CHUNK_TOKENS   = 50     # merge chunks smaller than this
MAX_CHUNKS_PER_FILE = 200   # if exceeded, merge more aggressively
MIN_FILE_CHARS     = 50     # filter files with less actual text than this
TOKEN_RATIO        = 1.3    # approx tokens per word for English technical text
NO_FILTER          = False  # if True, keep ALL files (no filtering)


# --- Data Structures ---------------------------------------------------------
@dataclass
class Chunk:
    content: str
    source_file: str
    page_title: str               # = filename without extension (TWiki page name)
    section_heading: str = ""
    chunk_index: int = 0
    total_chunks: int = 0
    author: str = ""
    date: str = ""
    content_type: str = "technical"  # technical | qa | meeting_log | table | minimal
    char_count: int = 0
    estimated_tokens: int = 0


@dataclass
class AnalysisStats:
    total_files: int = 0
    filtered_files: int = 0
    chunked_files: int = 0
    total_chunks: int = 0
    filter_reasons: dict = field(default_factory=dict)
    content_types: dict = field(default_factory=dict)
    chunk_sizes: list = field(default_factory=list)
    files_no_headings: int = 0
    files_with_html: int = 0
    files_with_qa: int = 0
    largest_file: str = ""
    largest_file_chunks: int = 0


# --- Utility Functions -------------------------------------------------------

def estimate_tokens(text):
    """Estimate token count. ~1.3 tokens per whitespace-delimited word."""
    words = text.split()
    return int(len(words) * TOKEN_RATIO)


def extract_signature(text):
    """
    Extract author and date from TWiki signature lines.
    Returns (cleaned_text, author, date).

    Patterns:
      -- AuthorName, 23 Jul 2009
      -- AuthorName, 2015-02-08
      -- AuthorName - 2023
    """
    author = ""
    date = ""

    sig_pattern = re.compile(
        r'^(?:\u2014|--)\s*'                # em-dash or double-dash
        r'([A-Za-z][A-Za-z0-9_ ]*?)'       # author name
        r'(?:\s*[-,]\s*'                    # separator
        r'(.+?))?'                          # optional date
        r'\s*$',
        re.MULTILINE
    )

    matches = list(sig_pattern.finditer(text))
    if matches:
        last = matches[-1]
        author = last.group(1).strip()
        date = (last.group(2) or "").strip()

    # Remove ALL signature lines from text
    cleaned = sig_pattern.sub('', text)

    return cleaned, author, date


# --- Phase 2: Cleaning ------------------------------------------------------

def clean_text(text):
    """Remove HTML tags, TWiki remnants, and conversion artifacts."""

    # Normalize line endings
    text = text.replace('\r\n', '\n').replace('\r', '\n')

    # Remove <sticky> tags
    text = re.sub(r'</?sticky>', '', text)

    # Remove TWiki remnants
    text = re.sub(r'["\s]*hideimg(?:right|left)=["\s]*["}%]*', '', text)
    text = re.sub(r'["\s]*start="hide\s*"["\s}%]*', '', text)
    text = re.sub(r'^["\s]*\}%\s*$', '', text, flags=re.MULTILINE)

    # Remove HTML comments
    text = re.sub(r'<!--.*?-->', '', text, flags=re.DOTALL)

    # Convert <br> to newlines
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)

    # Remove <center> tags
    text = re.sub(r'</?center>', '', text, flags=re.IGNORECASE)

    # Extract text from <p>, <span>, <a> tags (keep inner text)
    text = re.sub(r'<p[^>]*>(.*?)</p>', r'\1', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<span[^>]*>(.*?)</span>', r'\1', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<a[^>]*>(.*?)</a>', r'\1', text, flags=re.DOTALL | re.IGNORECASE)

    # Remove <img> tags entirely
    text = re.sub(r'<img[^>]*/?\s*>', '', text, flags=re.IGNORECASE)

    # Remove remaining HTML tags but keep content
    text = re.sub(r'<[^>]+>', '', text)

    # Decode common HTML entities
    html_entities = {
        '&rarr;': '->', '&larr;': '<-', '&radic;': 'sqrt',
        '&mu;': 'mu', '&ge;': '>=', '&le;': '<=',
        '&gt;': '>', '&lt;': '<', '&amp;': '&',
        '&nbsp;': ' ',
    }
    for entity, replacement in html_entities.items():
        text = text.replace(entity, replacement)
    # Catch remaining numeric entities
    text = re.sub(r'&#x?[0-9a-fA-F]+;', '', text)

    # Remove role="presentation" LaTeX remnants
    text = re.sub(r'" role="presentation">[^<]*', '', text)

    # Collapse multiple blank lines
    text = re.sub(r'\n{3,}', '\n\n', text)

    # Strip trailing whitespace per line
    lines = [line.rstrip() for line in text.split('\n')]
    text = '\n'.join(lines)

    return text.strip()


# --- Phase 1: Filtering -----------------------------------------------------

def get_text_content(text):
    """Strip all markup/formatting to get raw text for length check."""
    t = clean_text(text)
    t = re.sub(r'^#+\s*', '', t, flags=re.MULTILINE)
    t = re.sub(r'[*_`~\[\](){}|#>!]', '', t)
    t = re.sub(r'https?://\S+', '', t)
    t = re.sub(r'[\U0001f534\U0001f7e2\U0001f535\u2139\ufe0f\u2705\u2714\u274c]', '', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def should_filter(text, filename):
    """
    Decide if file should be filtered out.
    Returns reason string if filtered, None if kept.
    """
    raw_text = get_text_content(text)

    if len(raw_text) < MIN_FILE_CHARS:
        return "insufficient_content ({} chars)".format(len(raw_text))

    # Check if mostly just links with no descriptive text
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    if len(lines) > 10:
        link_lines = 0
        for l in lines:
            if re.search(r'https?://\S+', l):
                without_url = re.sub(r'https?://\S+', '', l).strip()
                # Line is "link-only" if removing URL leaves < 30 chars of non-markup
                clean_remainder = re.sub(r'[*_`\[\]()#>!\-|]', '', without_url).strip()
                if len(clean_remainder) < 30:
                    link_lines += 1
        link_ratio = link_lines / len(lines)
        if link_ratio > 0.7:
            return "mostly_links ({:.0%} link-only lines)".format(link_ratio)

    return None


# --- Phase 3: Chunking Strategies -------------------------------------------

def detect_content_type(text):
    """Classify content type based on patterns."""
    qa_count = len(re.findall(r'\U0001f534', text)) + len(re.findall(r'\U0001f7e2', text))
    if qa_count >= 4:
        return "qa"

    lines = [l.strip() for l in text.split('\n') if l.strip()]
    if len(lines) > 10:
        link_lines = sum(1 for l in lines if re.search(r'https?://\S+', l))
        if link_lines / len(lines) > 0.5:
            return "meeting_log"

    table_lines = sum(1 for l in lines if l.startswith('|'))
    if len(lines) > 0 and table_lines / len(lines) > 0.6:
        return "table"

    if len(get_text_content(text)) < 200:
        return "minimal"

    return "technical"


def split_by_headings(text):
    """
    Split text by markdown headings.
    Returns list of (heading, content) tuples.
    """
    heading_pattern = re.compile(r'^(#{1,6})\s+(.+)$', re.MULTILINE)

    sections = []
    last_end = 0
    last_heading = ""

    for match in heading_pattern.finditer(text):
        content_before = text[last_end:match.start()].strip()
        if content_before or last_heading:
            sections.append((last_heading, content_before))
        last_heading = match.group(2).strip()
        last_end = match.end()

    # Last section
    remaining = text[last_end:].strip()
    if remaining or last_heading:
        sections.append((last_heading, remaining))

    if not sections:
        sections = [("", text.strip())]

    return sections


def split_qa_pairs(text):
    """
    Split Q&A content into question-answer pairs.
    Groups consecutive red-circle (question) and green-circle (answer) blocks.
    """
    lines = text.split('\n')
    pairs = []
    current_block = []
    current_heading = ""
    in_qa = False

    for line in lines:
        heading_match = re.match(r'^(#{1,6})\s+(.+)$', line)
        if heading_match:
            if current_block:
                pairs.append((current_heading, '\n'.join(current_block)))
                current_block = []
            current_heading = heading_match.group(2).strip()
            in_qa = False
            continue

        is_question = bool(re.search(r'\U0001f534', line))
        is_answer = bool(re.search(r'\U0001f7e2', line))

        if is_question:
            if in_qa and current_block:
                pairs.append((current_heading, '\n'.join(current_block)))
                current_block = []
            in_qa = True

        if in_qa or is_question or is_answer:
            current_block.append(line)
        else:
            current_block.append(line)

    if current_block:
        pairs.append((current_heading, '\n'.join(current_block)))

    return pairs if pairs else [("", text)]


def split_by_paragraphs(text, max_tokens=None):
    """
    Recursive paragraph splitting for text without headings.
    """
    if max_tokens is None:
        max_tokens = MAX_CHUNK_TOKENS

    if estimate_tokens(text) <= max_tokens:
        return [text] if text.strip() else []

    # Try double newlines first
    paragraphs = re.split(r'\n\n+', text)
    if len(paragraphs) > 1:
        return _merge_splits(paragraphs, max_tokens)

    # Try single newlines
    lines = text.split('\n')
    if len(lines) > 1:
        return _merge_splits(lines, max_tokens)

    # Sentence boundaries
    sentences = re.split(r'(?<=[.!?])\s+', text)
    if len(sentences) > 1:
        return _merge_splits(sentences, max_tokens)

    # Hard split by words
    words = text.split()
    max_words = int(max_tokens / TOKEN_RATIO)
    chunks = []
    for i in range(0, len(words), max_words):
        chunk = ' '.join(words[i:i + max_words])
        if chunk:
            chunks.append(chunk)
    return chunks


def _merge_splits(parts, max_tokens):
    """Merge small consecutive parts until they approach max_tokens."""
    chunks = []
    current = ""

    for part in parts:
        part = part.strip()
        if not part:
            continue

        candidate = (current + '\n\n' + part).strip() if current else part

        if estimate_tokens(candidate) <= max_tokens:
            current = candidate
        else:
            if current:
                chunks.append(current)
            if estimate_tokens(part) > max_tokens:
                sub_chunks = split_by_paragraphs(part, max_tokens)
                chunks.extend(sub_chunks)
                current = ""
            else:
                current = part

    if current:
        chunks.append(current)

    return chunks


def add_overlap(chunks, overlap_tokens=None):
    """Add overlap between consecutive chunks by prepending tail of previous."""
    if overlap_tokens is None:
        overlap_tokens = OVERLAP_TOKENS

    if len(chunks) <= 1 or overlap_tokens <= 0:
        return chunks

    result = [chunks[0]]
    overlap_words = int(overlap_tokens / TOKEN_RATIO)

    for i in range(1, len(chunks)):
        prev_words = chunks[i - 1].split()
        if len(prev_words) > overlap_words:
            overlap_text = ' '.join(prev_words[-overlap_words:])
            result.append(overlap_text + '\n\n' + chunks[i])
        else:
            result.append(chunks[i])

    return result


# --- Main Chunking Logic ----------------------------------------------------

def chunk_file(filepath):
    """Process a single file through all three phases."""

    filename = Path(filepath).stem
    page_title = filename

    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        raw_text = f.read()

    # Phase 1: Filter (skip if NO_FILTER is True)
    if not NO_FILTER:
        filter_reason = should_filter(raw_text, filename)
        if filter_reason:
            return []

    # Phase 2: Clean
    cleaned, author, date = extract_signature(raw_text)
    cleaned = clean_text(cleaned)

    if not cleaned.strip():
        # Even with no-filter, truly empty files produce nothing
        return []

    # Detect content type
    content_type = detect_content_type(cleaned)

    # Phase 3: Chunk
    raw_chunks = []  # list of (heading, content)

    if content_type == "qa":
        qa_pairs = split_qa_pairs(cleaned)
        # Sub-split any Q&A blocks that are still too large
        for heading, content in qa_pairs:
            if not content.strip():
                continue
            tokens = estimate_tokens(content)
            if tokens <= MAX_CHUNK_TOKENS:
                raw_chunks.append((heading, content))
            else:
                sub_chunks = split_by_paragraphs(content, MAX_CHUNK_TOKENS)
                for i, sc in enumerate(sub_chunks):
                    sub_heading = heading if i == 0 else (heading + " (cont.)" if heading else "")
                    raw_chunks.append((sub_heading, sc))
    else:
        sections = split_by_headings(cleaned)
        for heading, content in sections:
            if not content.strip():
                continue
            tokens = estimate_tokens(content)
            if tokens <= MAX_CHUNK_TOKENS:
                raw_chunks.append((heading, content))
            else:
                sub_chunks = split_by_paragraphs(content, MAX_CHUNK_TOKENS)
                for i, sc in enumerate(sub_chunks):
                    sub_heading = heading if i == 0 else (heading + " (cont.)" if heading else "")
                    raw_chunks.append((sub_heading, sc))

    # ── Merge tiny chunks (bidirectional) ──
    merged = merge_tiny_chunks(raw_chunks, MIN_CHUNK_TOKENS)

    # ── Add overlap ──
    contents = [c for _, c in merged]
    headings = [h for h, _ in merged]
    overlapped = add_overlap(contents, OVERLAP_TOKENS)

    # ── Hard enforcement: nothing over 512 tokens ──
    final_contents = []
    final_headings = []
    for i, content in enumerate(overlapped):
        heading = headings[i] if i < len(headings) else ""
        if estimate_tokens(content) > 512:
            # Force-split by words
            words = content.split()
            max_words = int(512 / TOKEN_RATIO)
            for j in range(0, len(words), max_words):
                sub = ' '.join(words[j:j + max_words])
                if sub.strip():
                    sub_heading = heading if j == 0 else (heading + " (cont.)" if heading else "")
                    final_contents.append(sub)
                    final_headings.append(sub_heading)
        else:
            final_contents.append(content)
            final_headings.append(heading)

    # ── Build Chunk objects ──
    chunks = []
    total = len(final_contents)

    for i, content in enumerate(final_contents):
        content = content.strip()
        if not content:
            continue

        chunk = Chunk(
            content=content,
            source_file=filename,
            page_title=page_title,
            section_heading=final_headings[i] if i < len(final_headings) else "",
            chunk_index=i,
            total_chunks=total,
            author=author,
            date=date,
            content_type=content_type,
            char_count=len(content),
            estimated_tokens=estimate_tokens(content),
        )
        chunks.append(chunk)

    return chunks


def merge_tiny_chunks(chunks_with_headings, min_tokens):
    """
    Merge chunks below min_tokens threshold.
    Tries backward first (merge with previous), then forward (merge with next).
    Runs multiple passes until stable.
    """
    merged = list(chunks_with_headings)

    # Remove empty chunks
    merged = [(h, c.strip()) for h, c in merged if c.strip()]

    changed = True
    max_passes = 5
    pass_num = 0

    while changed and pass_num < max_passes:
        changed = False
        pass_num += 1
        new_merged = []

        i = 0
        while i < len(merged):
            heading, content = merged[i]
            tokens = estimate_tokens(content)

            if tokens < min_tokens:
                if new_merged:
                    # Merge backward (with previous)
                    prev_h, prev_c = new_merged[-1]
                    combined = prev_c + '\n\n' + content
                    # Only merge if result stays under max
                    if estimate_tokens(combined) <= MAX_CHUNK_TOKENS:
                        new_merged[-1] = (prev_h, combined)
                        changed = True
                        i += 1
                        continue
                if i + 1 < len(merged):
                    # Merge forward (with next)
                    next_h, next_c = merged[i + 1]
                    combined = content + '\n\n' + next_c
                    if estimate_tokens(combined) <= MAX_CHUNK_TOKENS:
                        merged[i + 1] = (heading or next_h, combined)
                        changed = True
                        i += 1
                        continue

            new_merged.append((heading, content))
            i += 1

        merged = new_merged

    return merged


# --- Analysis Mode -----------------------------------------------------------

def analyze_corpus(input_dir):
    """Run analysis on entire corpus and print statistics."""
    stats = AnalysisStats()
    md_files = sorted(Path(input_dir).glob('*.md'))
    stats.total_files = len(md_files)

    for filepath in md_files:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            raw_text = f.read()

        filename = filepath.stem

        if not NO_FILTER:
            reason = should_filter(raw_text, filename)
            if reason:
                stats.filtered_files += 1
                category = reason.split('(')[0].strip()
                stats.filter_reasons[category] = stats.filter_reasons.get(category, 0) + 1
                continue

        cleaned = clean_text(raw_text)
        ctype = detect_content_type(cleaned)
        stats.content_types[ctype] = stats.content_types.get(ctype, 0) + 1

        if not re.search(r'^#{1,6}\s+', cleaned, re.MULTILINE):
            stats.files_no_headings += 1

        if re.search(r'<[a-z][^>]*>', cleaned, re.IGNORECASE):
            stats.files_with_html += 1

        if ctype == "qa":
            stats.files_with_qa += 1

        chunks = chunk_file(str(filepath))
        stats.chunked_files += 1
        stats.total_chunks += len(chunks)

        for chunk in chunks:
            stats.chunk_sizes.append(chunk.estimated_tokens)

        if len(chunks) > stats.largest_file_chunks:
            stats.largest_file_chunks = len(chunks)
            stats.largest_file = filename

    return stats


def print_analysis(stats):
    """Print formatted analysis report."""
    print("=" * 65)
    print("  CMS TWiki Chunker - Corpus Analysis Report")
    print("=" * 65)

    print("\n--- FILES ---")
    print("  Total .md files found:    {:,}".format(stats.total_files))
    print("  Filtered out:             {:,}".format(stats.filtered_files))
    print("  Chunked:                  {:,}".format(stats.chunked_files))
    print("  Total chunks produced:    {:,}".format(stats.total_chunks))

    if stats.filter_reasons:
        print("\n--- FILTER REASONS ---")
        for reason, count in sorted(stats.filter_reasons.items(), key=lambda x: -x[1]):
            print("  {:<40} {:>6,}".format(reason, count))

    if stats.content_types:
        print("\n--- CONTENT TYPES ---")
        for ctype, count in sorted(stats.content_types.items(), key=lambda x: -x[1]):
            print("  {:<40} {:>6,}".format(ctype, count))

    print("\n--- STRUCTURAL OBSERVATIONS ---")
    print("  Files with no headings:   {:,}".format(stats.files_no_headings))
    print("  Files with residual HTML: {:,}".format(stats.files_with_html))
    print("  Files with Q&A patterns:  {:,}".format(stats.files_with_qa))

    if stats.chunk_sizes:
        avg = sum(stats.chunk_sizes) / len(stats.chunk_sizes)
        sorted_sizes = sorted(stats.chunk_sizes)
        median = sorted_sizes[len(sorted_sizes) // 2]
        mn = min(stats.chunk_sizes)
        mx = max(stats.chunk_sizes)
        over_limit = sum(1 for s in stats.chunk_sizes if s > 512)

        print("\n--- CHUNK SIZE DISTRIBUTION (estimated tokens) ---")
        print("  Min:                      {:,}".format(mn))
        print("  Median:                   {:,}".format(median))
        print("  Mean:                     {:,.1f}".format(avg))
        print("  Max:                      {:,}".format(mx))
        print("  Over 512 tokens:          {:,} ({:.1%})".format(over_limit, over_limit/len(stats.chunk_sizes)))

        # Histogram buckets
        buckets = [(0, 50), (50, 100), (100, 200), (200, 300), (300, 400), (400, 512), (512, 99999)]
        print("\n  {:<20} {:>8} {}".format("Range", "Count", "Bar"))
        for lo, hi in buckets:
            count = sum(1 for s in stats.chunk_sizes if lo <= s < hi)
            bar_len = int(count / max(len(stats.chunk_sizes), 1) * 50)
            bar = '#' * bar_len
            label = "{:>4}-{:>4}".format(lo, hi if hi < 99999 else "inf")
            print("  {:<20} {:>8,} {}".format(label, count, bar))

    print("\n  Largest file: {} ({} chunks)".format(stats.largest_file, stats.largest_file_chunks))
    print("=" * 65)


# --- Sample Mode -------------------------------------------------------------

def sample_file(filepath):
    """Process a single file and display chunks for inspection."""
    chunks = chunk_file(filepath)
    filename = Path(filepath).name

    if not chunks:
        print("File was filtered out (insufficient content or mostly links).")
        return

    print("=" * 65)
    print("  File: {}".format(filename))
    print("  Page title: {}".format(chunks[0].page_title))
    print("  Author: {}".format(chunks[0].author or '(none)'))
    print("  Date: {}".format(chunks[0].date or '(none)'))
    print("  Content type: {}".format(chunks[0].content_type))
    print("  Total chunks: {}".format(len(chunks)))
    print("=" * 65)

    for chunk in chunks:
        print("\n" + "-" * 65)
        print("  Chunk {}/{}  |  Section: {}  |  ~{} tokens".format(
            chunk.chunk_index + 1, chunk.total_chunks,
            chunk.section_heading or '(none)',
            chunk.estimated_tokens))
        print("-" * 65)
        preview = chunk.content[:500]
        if len(chunk.content) > 500:
            preview += "\n  ... [{} more chars]".format(len(chunk.content) - 500)
        print(preview)

    print("\n" + "=" * 65)


# --- Excel Report Generation -------------------------------------------------

def generate_excel_report(report_path, total_files, filtered_count, total_chunks,
                          error_count, filtered_log, all_chunks_log, issues_log,
                          error_log, file_summary_log):
    """Generate an Excel report documenting the chunking process."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  openpyxl not installed. Install with: pip install openpyxl")
        print("  Skipping Excel report generation.")
        return

    wb = Workbook()

    # Sanitize strings for Excel (remove illegal characters)
    import re as _re
    _ILLEGAL_CHARS = _re.compile(
        r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
    )
    def safe(val):
        """Make a value safe for openpyxl cells."""
        if val is None:
            return ""
        if isinstance(val, str):
            return _ILLEGAL_CHARS.sub('', val)
        return val

    # Colors and styles
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    title_font = Font(name='Arial', bold=True, size=14)
    label_font = Font(name='Arial', bold=True, size=11)
    normal_font = Font(name='Arial', size=11)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def write_header_row(ws, row, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def auto_width(ws, min_width=10, max_width=60):
        for col in ws.columns:
            col_letter = col[0].column_letter
            lengths = []
            for cell in col:
                if cell.value:
                    lengths.append(len(str(cell.value)))
            if lengths:
                width = min(max(max(lengths), min_width), max_width)
                ws.column_dimensions[col_letter].width = width + 2

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = "2F5496"

    ws_summary.cell(row=1, column=1, value="CMS TWiki Chunker Report").font = title_font
    ws_summary.cell(row=2, column=1, value="Generated by chunker.py").font = normal_font

    summary_data = [
        ("Total .md files scanned", total_files),
        ("Files filtered out", filtered_count),
        ("Files chunked", total_files - filtered_count - error_count),
        ("Total chunks produced", total_chunks),
        ("Processing errors", error_count),
        ("", ""),
        ("Settings", ""),
        ("Max tokens per chunk", MAX_CHUNK_TOKENS),
        ("Overlap tokens", OVERLAP_TOKENS),
        ("Min file chars threshold", MIN_FILE_CHARS),
    ]

    for i, (label, value) in enumerate(summary_data, 4):
        ws_summary.cell(row=i, column=1, value=label).font = label_font
        cell = ws_summary.cell(row=i, column=2, value=value)
        cell.font = normal_font
        if label == "Files filtered out" and value > 0:
            cell.fill = yellow_fill
        elif label == "Processing errors" and value > 0:
            cell.fill = red_fill
        elif label == "Total chunks produced":
            cell.fill = green_fill

    # Content type breakdown
    row = 16
    ws_summary.cell(row=row, column=1, value="Content Type Breakdown").font = label_font
    row += 1
    type_counts = {}
    for rec in all_chunks_log:
        ct = rec.get('content_type', 'unknown')
        type_counts[ct] = type_counts.get(ct, 0) + 1
    for ct, count in sorted(type_counts.items(), key=lambda x: -x[1]):
        ws_summary.cell(row=row, column=1, value=ct).font = normal_font
        ws_summary.cell(row=row, column=2, value=count).font = normal_font
        row += 1

    # Token distribution
    row += 1
    ws_summary.cell(row=row, column=1, value="Chunk Size Distribution").font = label_font
    row += 1
    if all_chunks_log:
        sizes = [r['estimated_tokens'] for r in all_chunks_log]
        buckets = [(0, 50), (50, 100), (100, 200), (200, 300), (300, 400), (400, 512)]
        for lo, hi in buckets:
            count = sum(1 for s in sizes if lo <= s < hi)
            ws_summary.cell(row=row, column=1, value="{}-{} tokens".format(lo, hi)).font = normal_font
            ws_summary.cell(row=row, column=2, value=count).font = normal_font
            row += 1

    auto_width(ws_summary)

    # ── Sheet 2: Filtered Files ──
    ws_filtered = wb.create_sheet("Filtered Files")
    ws_filtered.sheet_properties.tabColor = "FF6600"

    headers = ["Filename", "File Size (bytes)", "Filter Reason"]
    write_header_row(ws_filtered, 1, headers)

    for i, (filename, size, reason) in enumerate(filtered_log, 2):
        ws_filtered.cell(row=i, column=1, value=safe(filename)).font = normal_font
        ws_filtered.cell(row=i, column=2, value=size).font = normal_font
        ws_filtered.cell(row=i, column=3, value=safe(reason)).font = normal_font
        for col in range(1, 4):
            ws_filtered.cell(row=i, column=col).border = thin_border

    if not filtered_log:
        ws_filtered.cell(row=2, column=1, value="(no files were filtered)").font = normal_font

    auto_width(ws_filtered)

    # ── Sheet 3: File Summary ──
    ws_files = wb.create_sheet("File Summary")
    ws_files.sheet_properties.tabColor = "00B050"

    headers = ["Filename", "Size (bytes)", "Content Type", "Num Chunks", "Author", "Date"]
    write_header_row(ws_files, 1, headers)

    for i, (filename, size, ctype, num_chunks, author, date) in enumerate(file_summary_log, 2):
        ws_files.cell(row=i, column=1, value=safe(filename)).font = normal_font
        ws_files.cell(row=i, column=2, value=size).font = normal_font
        ws_files.cell(row=i, column=3, value=safe(ctype)).font = normal_font
        ws_files.cell(row=i, column=4, value=num_chunks).font = normal_font
        ws_files.cell(row=i, column=5, value=safe(author)).font = normal_font
        ws_files.cell(row=i, column=6, value=safe(date)).font = normal_font
        for col in range(1, 7):
            ws_files.cell(row=i, column=col).border = thin_border

    auto_width(ws_files)

    # ── Sheet 4: All Chunks ──
    ws_chunks = wb.create_sheet("All Chunks")
    ws_chunks.sheet_properties.tabColor = "4472C4"

    headers = ["Source File", "Chunk #", "Total Chunks", "Section Heading",
               "Content Type", "Est. Tokens", "Char Count", "Author",
               "Content Preview (first 200 chars)"]
    write_header_row(ws_chunks, 1, headers)

    for i, rec in enumerate(all_chunks_log, 2):
        ws_chunks.cell(row=i, column=1, value=safe(rec['source_file'])).font = normal_font
        ws_chunks.cell(row=i, column=2, value=rec['chunk_index'] + 1).font = normal_font
        ws_chunks.cell(row=i, column=3, value=rec['total_chunks']).font = normal_font
        ws_chunks.cell(row=i, column=4, value=safe(rec['section_heading'])).font = normal_font
        ws_chunks.cell(row=i, column=5, value=safe(rec['content_type'])).font = normal_font
        ws_chunks.cell(row=i, column=6, value=rec['estimated_tokens']).font = normal_font
        ws_chunks.cell(row=i, column=7, value=rec['char_count']).font = normal_font
        ws_chunks.cell(row=i, column=8, value=safe(rec['author'])).font = normal_font
        preview = rec['content'][:200].replace('\n', ' ')
        ws_chunks.cell(row=i, column=9, value=safe(preview)).font = normal_font
        for col in range(1, 10):
            ws_chunks.cell(row=i, column=col).border = thin_border
        # Highlight token issues
        if rec['estimated_tokens'] > 450:
            ws_chunks.cell(row=i, column=6).fill = yellow_fill
        elif rec['estimated_tokens'] < 10:
            ws_chunks.cell(row=i, column=6).fill = red_fill

    auto_width(ws_chunks)

    # ── Sheet 5: Issues ──
    ws_issues = wb.create_sheet("Issues")
    ws_issues.sheet_properties.tabColor = "FF0000"

    headers = ["Filename", "Issue Type", "Details"]
    write_header_row(ws_issues, 1, headers)

    for i, (filename, issue_type, details) in enumerate(issues_log, 2):
        ws_issues.cell(row=i, column=1, value=safe(filename)).font = normal_font
        ws_issues.cell(row=i, column=2, value=safe(issue_type)).font = normal_font
        ws_issues.cell(row=i, column=3, value=safe(details)).font = normal_font
        for col in range(1, 4):
            ws_issues.cell(row=i, column=col).border = thin_border
        # Color by severity
        if issue_type in ('residual_html', 'large_chunk'):
            ws_issues.cell(row=i, column=2).fill = yellow_fill
        elif issue_type == 'tiny_chunk':
            ws_issues.cell(row=i, column=2).fill = red_fill
        elif issue_type == 'no_headings':
            ws_issues.cell(row=i, column=2).fill = PatternFill(
                start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

    if not issues_log:
        ws_issues.cell(row=2, column=1, value="(no issues detected)").font = normal_font

    auto_width(ws_issues)

    # ── Sheet 6: Errors (if any) ──
    if error_log:
        ws_errors = wb.create_sheet("Errors")
        ws_errors.sheet_properties.tabColor = "C00000"
        headers = ["Filename", "Error Message"]
        write_header_row(ws_errors, 1, headers)
        for i, (filename, msg) in enumerate(error_log, 2):
            ws_errors.cell(row=i, column=1, value=safe(filename)).font = normal_font
            ws_errors.cell(row=i, column=2, value=safe(msg)).font = normal_font
        auto_width(ws_errors)

    # Save
    wb.save(report_path)
    print("  Report saved: {}".format(report_path))


def _apply_config(max_tokens, overlap, min_chars, no_filter=False):
    """Update module-level configuration."""
    global MAX_CHUNK_TOKENS, OVERLAP_TOKENS, MIN_FILE_CHARS, MAX_CHUNKS_PER_FILE, NO_FILTER
    MAX_CHUNK_TOKENS = max_tokens
    OVERLAP_TOKENS = overlap
    MIN_FILE_CHARS = min_chars
    NO_FILTER = no_filter


# --- Main Entry Point --------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='CMS TWiki GFM Chunker for RAG Pipeline',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python chunker.py ./md_files output.jsonl              Chunk all files
  python chunker.py ./md_files output.jsonl --analyze    Analyze corpus first
  python chunker.py --sample myfile.md                   Inspect one file
        """
    )
    parser.add_argument('input_dir', nargs='?', help='Directory of .md files')
    parser.add_argument('output_file', nargs='?', help='Output .jsonl file')
    parser.add_argument('--analyze', action='store_true',
                        help='Run analysis mode (stats only, no output file)')
    parser.add_argument('--sample', type=str,
                        help='Process a single file and display chunks')
    parser.add_argument('--max-tokens', type=int, default=MAX_CHUNK_TOKENS,
                        help='Max tokens per chunk (default: {})'.format(MAX_CHUNK_TOKENS))
    parser.add_argument('--overlap', type=int, default=OVERLAP_TOKENS,
                        help='Overlap tokens (default: {})'.format(OVERLAP_TOKENS))
    parser.add_argument('--min-chars', type=int, default=MIN_FILE_CHARS,
                        help='Min chars to keep file (default: {})'.format(MIN_FILE_CHARS))
    parser.add_argument('--no-filter', action='store_true',
                        help='Keep ALL files, no filtering (for completeness)')

    args = parser.parse_args()

    # Update module-level config if custom values given
    _apply_config(args.max_tokens, args.overlap, args.min_chars, args.no_filter)

    # Sample mode
    if args.sample:
        if not os.path.isfile(args.sample):
            print("Error: File not found: {}".format(args.sample))
            sys.exit(1)
        sample_file(args.sample)
        return

    if not args.input_dir:
        parser.error("input_dir is required (unless using --sample)")

    if not os.path.isdir(args.input_dir):
        print("Error: Directory not found: {}".format(args.input_dir))
        sys.exit(1)

    # Analysis mode
    if args.analyze:
        print("Analyzing corpus...")
        stats = analyze_corpus(args.input_dir)
        print_analysis(stats)
        return

    # Full chunking mode
    if not args.output_file:
        parser.error("output_file is required for chunking mode")

    md_files = sorted(Path(args.input_dir).glob('*.md'))
    total_files = len(md_files)

    if total_files == 0:
        print("No .md files found in {}".format(args.input_dir))
        sys.exit(1)

    print("Processing {:,} files from {}".format(total_files, args.input_dir))
    print("Settings: max_tokens={}, overlap={}, min_chars={}".format(
        MAX_CHUNK_TOKENS, OVERLAP_TOKENS, MIN_FILE_CHARS))

    # Collect data for the Excel report
    filtered_log = []       # (filename, file_size_bytes, reason)
    all_chunks_log = []     # list of chunk dicts
    issues_log = []         # (filename, issue_type, details)
    error_log = []          # (filename, error_message)
    file_summary_log = []   # (filename, file_size, content_type, num_chunks, author, date)

    filtered_count = 0
    total_chunks = 0
    error_count = 0

    with open(args.output_file, 'w', encoding='utf-8') as out:
        for i, filepath in enumerate(md_files):
            if (i + 1) % 5000 == 0 or i == 0:
                print("  [{:>6,}/{:,}] Processing {}...".format(i + 1, total_files, filepath.name))

            try:
                file_size = filepath.stat().st_size
                filename = filepath.name

                # Check filter before chunking (to log reason)
                with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                    raw_text = f.read()

                if not NO_FILTER:
                    filter_reason = should_filter(raw_text, filepath.stem)
                    if filter_reason:
                        filtered_count += 1
                        filtered_log.append((filename, file_size, filter_reason))
                        continue

                chunks = chunk_file(str(filepath))
                if not chunks:
                    filtered_count += 1
                    filtered_log.append((filename, file_size, "empty_after_cleaning"))
                    continue

                # Log file summary
                first = chunks[0]
                file_summary_log.append((
                    filename, file_size, first.content_type,
                    len(chunks), first.author, first.date
                ))

                # Check for issues
                cleaned = clean_text(raw_text)
                if not re.search(r'^#{1,6}\s+', cleaned, re.MULTILINE):
                    issues_log.append((filename, "no_headings",
                        "File has no markdown headings; chunked by paragraph/fallback"))

                if re.search(r'<[a-z][^>]*>', cleaned, re.IGNORECASE):
                    residual = re.findall(r'<[a-z][^>]*>', cleaned, re.IGNORECASE)
                    issues_log.append((filename, "residual_html",
                        "{} HTML tags remain after cleaning".format(len(residual))))

                for chunk in chunks:
                    record = asdict(chunk)
                    out.write(json.dumps(record, ensure_ascii=False) + '\n')
                    total_chunks += 1
                    all_chunks_log.append(record)

                    # Flag problematic chunks
                    if chunk.estimated_tokens < 10:
                        issues_log.append((filename, "tiny_chunk",
                            "Chunk {}/{} has only ~{} tokens".format(
                                chunk.chunk_index + 1, chunk.total_chunks,
                                chunk.estimated_tokens)))
                    if chunk.estimated_tokens > 450:
                        issues_log.append((filename, "large_chunk",
                            "Chunk {}/{} has ~{} tokens (close to 512 limit)".format(
                                chunk.chunk_index + 1, chunk.total_chunks,
                                chunk.estimated_tokens)))

            except Exception as e:
                error_count += 1
                error_log.append((filepath.name, str(e)))
                print("  Error processing {}: {}".format(filepath.name, e))

    print("\nDone!")
    print("  Files processed: {:,}".format(total_files - filtered_count))
    print("  Files filtered:  {:,}".format(filtered_count))
    print("  Total chunks:    {:,}".format(total_chunks))
    print("  Errors:          {:,}".format(error_count))
    print("  Output:          {}".format(args.output_file))

    # Generate Excel report
    report_path = args.output_file.replace('.jsonl', '_report.xlsx')
    if report_path == args.output_file:
        report_path = args.output_file + '_report.xlsx'

    print("  Generating report: {}".format(report_path))
    generate_excel_report(
        report_path, total_files, filtered_count, total_chunks, error_count,
        filtered_log, all_chunks_log, issues_log, error_log, file_summary_log
    )


if __name__ == '__main__':
    main()
