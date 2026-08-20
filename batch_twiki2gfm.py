"""
TWiki2GFM Batch Converter + Excel Report
==========================================
Runs the twiki2gfm converter on all TWiki files and generates
a detailed Excel report of results.

Usage:
    python batch_twiki2gfm.py

Expects to be in the same directory as twiki2gfm.py
"""

import os
import re
import sys
import time
from pathlib import Path

# ── Make sure twiki2gfm is importable ──
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from twiki2gfm import TWikiConverter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Configuration ──
INPUT_DIR = r"pandoc_convert\Raw_Twiki_Files\raw"
OUTPUT_DIR = "TWIKI2GFM_OUTPUT"
REPORT_FILE = "twiki2gfm_report.xlsx"


def detect_applied_rules(original: str, converted: str) -> list:
    """Detect which conversion rules were actually applied."""
    rules = []

    if re.search(r'^%META:', original, re.MULTILINE):
        if not re.search(r'^%META:', converted, re.MULTILINE):
            rules.append("META stripped")

    if re.search(r'^-{3,}\s*\++', original, re.MULTILINE):
        if re.search(r'^#{1,6}\s', converted, re.MULTILINE):
            rules.append("Headings converted")

    if re.search(r'<verbatim', original):
        if '```' in converted:
            rules.append("Verbatim → code blocks")

    if re.search(r'\[\[.*?\]\]', original):
        rules.append("Links converted")

    if re.search(r'^\s*\|.*\|\s*$', original, re.MULTILINE):
        if re.search(r'\|[-:]+\|', converted):
            rules.append("Tables converted")

    if re.search(r'^\s{3,}\*\s', original, re.MULTILINE):
        if re.search(r'^(\s*)- ', converted, re.MULTILINE):
            rules.append("Bullet lists converted")

    if re.search(r'^\s{3,}\d+\.\s', original, re.MULTILINE):
        rules.append("Numbered lists converted")

    color_pattern = '|'.join([
        'RED', 'GREEN', 'BLUE', 'YELLOW', 'ORANGE', 'PURPLE',
        'PINK', 'AQUA', 'MAROON', 'NAVY', 'TEAL', 'LIME',
        'OLIVE', 'SILVER', 'GRAY', 'BLACK', 'WHITE',
    ])
    if re.search(rf'%({color_pattern})%', original):
        rules.append("Colors stripped")

    if re.search(r'%[A-Z_]+(\{[^}]*\})?%', original):
        rules.append("Variables processed")

    if re.search(r'(?<!\*)\*(?!\*)(\S.*?\S|\S)\*(?!\*)', original):
        if '**' in converted:
            rules.append("Bold converted")

    if re.search(r'(?<!_)_(?!_)(\S.*?\S|\S)_(?!_)', original):
        rules.append("Italic converted")

    if re.search(r'(?<!=)=(?!=)(\S.*?\S|\S)=(?!=)', original):
        rules.append("Code spans converted")

    if re.search(r'<br\s*/?>', original):
        rules.append("BR tags normalized")

    if re.search(r'<nop>', original):
        rules.append("Nop tags stripped")

    if re.search(r'<noautolink>', original):
        rules.append("Noautolink stripped")

    if re.search(r'-- (Main\.|TWiki\.)', original):
        rules.append("Signatures converted")

    if re.search(r'!\w+[A-Z]\w+', original):
        rules.append("WikiWord escapes stripped")

    return rules if rules else ["No specific rules matched"]


def main():
    input_path = Path(INPUT_DIR)
    output_path = Path(OUTPUT_DIR)
    output_path.mkdir(parents=True, exist_ok=True)

    # Find all .txt files recursively
    all_files = sorted(input_path.rglob("*.txt"), key=lambda p: p.name.casefold())
    total = len(all_files)

    if total == 0:
        print(f"No .txt files found in {INPUT_DIR}")
        sys.exit(1)

    print(f"Found {total} files to convert")
    print(f"Output: {OUTPUT_DIR}/")
    print(f"{'=' * 60}")

    converter = TWikiConverter()
    results = []
    success = 0
    warnings_count = 0
    errors_count = 0
    start_time = time.time()

    for idx, fpath in enumerate(all_files, 1):
        basename = fpath.stem
        out_file = output_path / f"{basename}.md"

        try:
            # Read input
            with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
                original_text = f.read()

            # Convert
            converter.verbatim_blocks = []
            converter.warnings = []
            converted_text = converter.convert(original_text)

            # Validate
            validation = converter.validate(converted_text)

            # Detect applied rules
            applied_rules = detect_applied_rules(original_text, converted_text)

            # Write output
            with open(out_file, 'w', encoding='utf-8') as f:
                f.write(converted_text)

            # Determine status
            has_errors = any('❌' in v for v in validation)
            has_warnings = any('⚠️' in v for v in validation)

            if has_errors:
                status = "Error"
                errors_count += 1
            elif has_warnings:
                status = "Warning"
                warnings_count += 1
            else:
                status = "Success"
                success += 1

            results.append({
                "input_name": fpath.name,
                "output_name": out_file.name,
                "status": status,
                "input_size": fpath.stat().st_size,
                "output_size": out_file.stat().st_size,
                "applied_rules": "; ".join(applied_rules),
                "validation": "; ".join(validation),
                "converter_warnings": "; ".join(converter.warnings) if converter.warnings else "None",
                "subfolder": str(fpath.parent.relative_to(input_path)) if fpath.parent != input_path else "root",
            })

        except Exception as e:
            errors_count += 1
            results.append({
                "input_name": fpath.name,
                "output_name": "—",
                "status": "Exception",
                "input_size": fpath.stat().st_size if fpath.exists() else 0,
                "output_size": 0,
                "applied_rules": "—",
                "validation": f"Exception: {str(e)}",
                "converter_warnings": "—",
                "subfolder": str(fpath.parent.relative_to(input_path)) if fpath.parent != input_path else "root",
            })

        # Progress every 500 files
        if idx % 500 == 0:
            elapsed = time.time() - start_time
            rate = idx / elapsed
            eta = (total - idx) / rate / 60
            print(f"  [{idx}/{total}]  ok={success} warn={warnings_count} err={errors_count}  ({rate:.0f} files/sec, ~{eta:.0f} min left)")

    elapsed = time.time() - start_time
    print(f"\n{'=' * 60}")
    print(f"  CONVERSION COMPLETE")
    print(f"{'=' * 60}")
    print(f"  Total:      {total}")
    print(f"  Success:    {success}")
    print(f"  Warnings:   {warnings_count}")
    print(f"  Errors:     {errors_count}")
    print(f"  Time:       {elapsed:.0f}s ({elapsed/60:.1f} min)")
    print(f"{'=' * 60}")

    # ── Generate Excel Report ──
    print(f"\nGenerating Excel report: {REPORT_FILE}")
    generate_excel_report(results, total, success, warnings_count, errors_count, elapsed)
    print(f"Done!")


def generate_excel_report(results, total, success, warnings_count, errors_count, elapsed):
    wb = Workbook()

    # Styles
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    title_font = Font(name="Arial", bold=True, size=14)
    label_font = Font(name="Arial", bold=True, size=11)
    value_font = Font(name="Arial", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "TWiki2GFM Batch Conversion Report"
    ws["A1"].font = title_font

    summary_data = [
        ("Total Input Files", total),
        ("Converted Successfully", success),
        ("Converted with Warnings", warnings_count),
        ("Errors / Exceptions", errors_count),
        ("Success Rate", f"{success / total * 100:.1f}%" if total else "0%"),
        ("Total Time", f"{elapsed:.0f}s ({elapsed/60:.1f} min)"),
        ("Converter", "twiki2gfm.py (15-phase regex pipeline)"),
    ]

    for i, (label, value) in enumerate(summary_data, 3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = label_font
        ws[f"A{i}"].border = thin_border
        ws[f"B{i}"] = value
        ws[f"B{i}"].font = value_font
        ws[f"B{i}"].border = thin_border

    ws["B3"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws["B5"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ws["B6"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Rule frequency
    ws["A12"] = "Most Common Applied Rules"
    ws["A12"].font = label_font
    rule_counts = {}
    for r in results:
        for rule in r["applied_rules"].split("; "):
            if rule and rule != "—":
                rule_counts[rule] = rule_counts.get(rule, 0) + 1

    sorted_rules = sorted(rule_counts.items(), key=lambda x: -x[1])
    for i, (rule, count) in enumerate(sorted_rules[:15], 13):
        ws[f"A{i}"] = rule
        ws[f"A{i}"].font = value_font
        ws[f"A{i}"].border = thin_border
        ws[f"B{i}"] = count
        ws[f"B{i}"].font = value_font
        ws[f"B{i}"].border = thin_border
        ws[f"C{i}"] = f"{count/total*100:.1f}%"
        ws[f"C{i}"].font = value_font
        ws[f"C{i}"].border = thin_border

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10

    # ── Sheet 2: All Files ──
    ws2 = wb.create_sheet("All Files")
    headers = ["#", "Input File", "Output File", "Status", "Input KB", "Output KB",
               "Applied Rules", "Validation", "Warnings", "Subfolder"]

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, r in enumerate(results, 2):
        ws2.cell(row=i, column=1, value=i-1).border = thin_border
        ws2.cell(row=i, column=2, value=r["input_name"]).border = thin_border
        ws2.cell(row=i, column=3, value=r["output_name"]).border = thin_border

        status_cell = ws2.cell(row=i, column=4, value=r["status"])
        status_cell.border = thin_border
        if r["status"] == "Success":
            status_cell.fill = success_fill
        elif r["status"] == "Warning":
            status_cell.fill = warning_fill
        else:
            status_cell.fill = error_fill

        ws2.cell(row=i, column=5, value=round(r["input_size"]/1024, 1)).border = thin_border
        ws2.cell(row=i, column=6, value=round(r["output_size"]/1024, 1)).border = thin_border
        ws2.cell(row=i, column=7, value=r["applied_rules"]).border = thin_border
        ws2.cell(row=i, column=8, value=r["validation"]).border = thin_border
        ws2.cell(row=i, column=9, value=r["converter_warnings"]).border = thin_border
        ws2.cell(row=i, column=10, value=r["subfolder"]).border = thin_border

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 50
    ws2.column_dimensions["H"].width = 45
    ws2.column_dimensions["I"].width = 30
    ws2.column_dimensions["J"].width = 25
    ws2.auto_filter.ref = f"A1:J{len(results)+1}"

    # ── Sheet 3: Warnings & Errors Only ──
    ws3 = wb.create_sheet("Issues")
    issue_headers = ["#", "File", "Status", "Validation", "Warnings"]
    for col, h in enumerate(issue_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.border = thin_border

    issue_rows = [r for r in results if r["status"] != "Success"]
    for i, r in enumerate(issue_rows, 2):
        ws3.cell(row=i, column=1, value=i-1).border = thin_border
        ws3.cell(row=i, column=2, value=r["input_name"]).border = thin_border
        ws3.cell(row=i, column=3, value=r["status"]).border = thin_border
        ws3.cell(row=i, column=4, value=r["validation"]).border = thin_border
        ws3.cell(row=i, column=5, value=r["converter_warnings"]).border = thin_border

    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 50
    ws3.column_dimensions["E"].width = 40
    ws3.auto_filter.ref = f"A1:E{len(issue_rows)+1}"

    # ── Sheet 4: Rule Frequency ──
    ws4 = wb.create_sheet("Rule Statistics")
    rule_headers = ["Rule", "Count", "% of Files"]
    for col, h in enumerate(rule_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, (rule, count) in enumerate(sorted_rules, 2):
        ws4.cell(row=i, column=1, value=rule).border = thin_border
        ws4.cell(row=i, column=2, value=count).border = thin_border
        ws4.cell(row=i, column=3, value=f"{count/total*100:.1f}%").border = thin_border

    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 12

    wb.save(REPORT_FILE)
    print(f"  Report saved: {REPORT_FILE}")


if __name__ == "__main__":
    main()
